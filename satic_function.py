import os
import random
import shutil
import sqlite3
import string
import time
import zipfile
from typing import Tuple

import cv2
import filetype
import imagehash
import natsort
import numpy
import rarfile
from PIL import Image
from openpyxl import Workbook, load_workbook, styles

temp_image_folder = 'temp_image_folder'
path_7zip = '7-Zip/7z.exe'
icon_del = r'icon/del.png'
icon_archive = r'icon/archive.png'
icon_folder = r'icon/folder.png'
icon_previous = r'icon/previous.png'
icon_recycle_bin = r'icon/recycle_bin.png'
icon_next = r'icon/next.png'
icon_refresh = r'icon/refresh.png'
history_back_dir = '查重结果'
cache_filename = 'hash_cache.db'
icon_next_5p = r'icon/next_5p.png'
icon_previous_5p = r'icon/previous_5p.png'


def print_function_info(mode: str = 'current'):
    """打印当前/上一个执行的函数信息
    传参：mode 'current' 或 'last'"""
    pass

    # if mode == 'current':
    #     print(time.strftime('%H:%M:%S ', time.localtime()),
    #           inspect.getframeinfo(inspect.currentframe().f_back).function)
    # elif mode == 'last':
    #     print(time.strftime('%H:%M:%S ', time.localtime()),
    #           inspect.getframeinfo(inspect.currentframe().f_back.f_back).function)


def merge_intersecting_tuples(tuples_list: list) -> list:
    """合并list中的有交集的元组 [(1,2),(2,3)]->[(1,2,3)]"""
    merged_list = []

    for i in range(len(tuples_list)):
        tuple_merged = False

        for j in range(len(merged_list)):
            if set(tuples_list[i]) & set(merged_list[j]):
                merged_list[j] = tuple(set(tuples_list[i]) | set(merged_list[j]))
                tuple_merged = True
                break

        if not tuple_merged:
            merged_list.append(tuples_list[i])

    return merged_list


def clear_temp_image_folder():
    """清空临时存放图片的文件夹"""
    print_function_info()
    if os.path.exists(temp_image_folder):
        for i in os.listdir(temp_image_folder):
            fullpath = os.path.join(temp_image_folder, i)
            if os.path.isfile(fullpath):
                os.remove(fullpath)
    else:
        os.mkdir(temp_image_folder)


def is_image(filepath, prec_mode=False):
    """检查传入路径文件是否为图片
    prec_mode决定是否采用filetype库查询"""
    print_function_info()
    if not os.path.exists(filepath):
        return False
    elif os.path.isdir(filepath):
        return False
    else:
        if prec_mode:
            return filetype.is_image(filepath)
        else:
            file_suffix = os.path.splitext(filepath)[1][1:].lower()  # 提取文件后缀名（不带.）
            image_suffix = ['jpg', 'png', 'webp']
            if file_suffix in image_suffix:
                return True
            else:
                return False


def is_archive(filepath, prec_mode=False):
    """检查传入路径文件是否为压缩包"""
    print_function_info()
    if not os.path.exists(filepath):
        return False
    elif os.path.isdir(filepath):
        return False
    else:
        if prec_mode:
            archive_type = ['zip', 'tar', 'rar', 'gz', '7z', 'xz']  # filetype库支持的压缩文件后缀名
            kind = filetype.guess(filepath)
            if kind is None:
                type_kind = None
            else:
                type_kind = kind.extension
            if type_kind in archive_type:
                return True
            else:
                return False
        else:
            file_suffix = os.path.splitext(filepath)[1][1:].lower()  # 提取文件后缀名（不带.）
            archive_suffix = ['zip', 'rar', '7z']
            if file_suffix in archive_suffix:
                return True
            else:
                return False


def get_folder_size(folder: str) -> int:
    """获取指定文件夹的总大小/byte
    传参：folder 文件夹路径str
    返回值：total_size 总大小int"""
    print_function_info()
    folder_size = 0
    for dirpath, dirnames, filenames in os.walk(folder):
        for item in filenames:
            filepath = os.path.join(dirpath, item)
            folder_size += os.path.getsize(filepath)
    return folder_size


def create_random_string(length: int):
    """生成一个指定长度的随机字符串（小写英文+数字）
    传参：length 字符串的长度
    返回值：生成的str"""
    print_function_info()
    characters = string.ascii_lowercase + string.digits
    random_string = ''.join(random.choices(characters, k=length))

    return random_string


'''替换为zipfile和rarfile，不再调用7zip解压
def extract_image_from_archive(filepath: str, extract_file_number=1):
    """解压压缩包中的指定数量的图片到指定文件夹"""
    print_function_info()
    # 提取压缩包内结构
    command_l = [path_7zip, "l", filepath]
    process_l = subprocess.run(command_l,
                               stdout=subprocess.PIPE,
                               creationflags=subprocess.CREATE_NO_WINDOW,
                               text=True,
                               universal_newlines=True)
    if process_l.returncode == 0:
        stdout_line = process_l.stdout.split('\n')
        # 提取出其中的图片文件列表
        """
        7zip stdout流输出示例
        Date      Time    Attr         Size   Compressed  Name
        ------------------- ----- ------------ ------------  ------------------------
        2023-11-03 15:49:27 ....A         6847         6847  1.xlsx
        """
        fileline = [i for i in stdout_line if i.find('....A') != -1]
        image_in_archive = []
        for line in fileline:
            split_type1 = line.split()
            find_start_text = split_type1[5]
            split_type2 = line.split(' ')
            find_start_index = split_type2.index(find_start_text)
            file_archive = ' '.join(split_type2[find_start_index:]).strip()
            for suffix in ['.jpg', '.png', '.webp']:
                if suffix in file_archive.lower():  # 筛选出图片文件
                    image_in_archive.append(file_archive)
        image_in_archive = natsort.natsorted(image_in_archive)
        # 解压对应图片
        if extract_file_number > len(image_in_archive):
            ex_number = len(image_in_archive)
        else:
            ex_number = extract_file_number

        extract_image_list = set()
        for index in range(ex_number):
            extract_image = image_in_archive[index]
            command_e = [path_7zip, "e", filepath, extract_image, '-o' + temp_image_folder]
            process_e = subprocess.run(command_e,
                                       stdout=subprocess.PIPE,
                                       creationflags=subprocess.CREATE_NO_WINDOW,
                                       text=True,
                                       universal_newlines=True)
            if process_e.returncode == 0:
                stdout_result = process_e.stdout
                if 'No files to process' in stdout_result:
                    pass  # 备忘录：7zip的stdout编码问题，中文系统gbk不支持特殊字符的解析，会被替换为_
                else:
                    image_name = extract_image.split('\\')[-1]
                    image_file = os.path.join(temp_image_folder, image_name)
                    suffix = os.path.splitext(image_file)[1]
                    new_name = f'{index}_' + create_random_string(16) + suffix
                    new_file = os.path.join(temp_image_folder, new_name)
                    os.rename(image_file, new_file)
                    extract_image_list.add(new_file)

        return extract_image_list, len(image_in_archive)
'''


def extract_image_from_archive(filepath: str, extract_file_number=1):
    """解压zip/rar压缩包中的指定数量的图片到指定文件夹"""
    try:
        archive_file = zipfile.ZipFile(filepath)
    except zipfile.BadZipFile:
        try:
            archive_file = rarfile.RarFile(filepath)
        except rarfile.NotRarFile:
            return [], 0  # 按照正常格式返回空值

    filelist = archive_file.namelist()  # 中文会变为乱码，需要转utf-8编码
    image_in_archive = []
    for file in filelist:
        for suffix in ['.jpg', '.png', '.webp']:
            if suffix in file.lower():
                image_in_archive.append(file)
    image_in_archive = natsort.natsorted(image_in_archive)

    if extract_file_number > len(image_in_archive):
        ex_number = len(image_in_archive)
    else:
        ex_number = extract_file_number

    extract_image_list = set()
    new_temp_image_folder = temp_image_folder + r'\temp'
    for index in range(ex_number):
        # 解压
        extract_image_path = image_in_archive[index]
        archive_file.extract(extract_image_path, new_temp_image_folder)
        # 改名
        local_image_file = os.path.join(new_temp_image_folder, extract_image_path)
        suffix = os.path.splitext(local_image_file)[1]
        new_name = f'{index}_' + create_random_string(16) + suffix
        new_file = os.path.join(temp_image_folder, new_name)
        # 移动
        shutil.move(local_image_file, new_file)
        extract_image_list.add(new_file)

    archive_file.close()
    # send2trash.send2trash(new_temp_image_folder)

    return extract_image_list, len(image_in_archive)


def convert_image_to_numpy(image_file, gray=False, resize: Tuple[int, int] = None):
    """将图片转换为numpy图片对象"""
    image_numpy = cv2.imdecode(numpy.fromfile(image_file, dtype=numpy.uint8), -1)

    if resize:
        image_numpy = cv2.resize(image_numpy, resize)

    if gray:
        try:
            image_numpy = cv2.cvtColor(image_numpy, cv2.COLOR_BGR2GRAY)
        except cv2.error:
            pass

    return image_numpy


def calc_images_ssim(image_1, image_2):
    """计算两张图片的ssim相似度"""
    print_function_info()
    image_1_numpy = convert_image_to_numpy(image_1, gray=True, resize=(8, 8))
    image_2_numpy = convert_image_to_numpy(image_2, gray=True, resize=(8, 8))

    # 计算均值、方差和协方差
    mean1, mean2 = numpy.mean(image_1_numpy), numpy.mean(image_2_numpy)
    var1, var2 = numpy.var(image_1_numpy), numpy.var(image_2_numpy)
    covar = numpy.cov(image_1_numpy.flatten(), image_2_numpy.flatten())[0][1]

    # 设置常数
    c1 = (0.01 * 255) ** 2
    c2 = (0.03 * 255) ** 2

    # 计算SSIM
    numerator = (2 * mean1 * mean2 + c1) * (2 * covar + c2)
    denominator = (mean1 ** 2 + mean2 ** 2 + c1) * (var1 + var2 + c2)
    ssim = numerator / denominator

    return ssim


def get_image_attr(imagefile, mode_hash: str):
    """计算图片的特征值
    传参 mode_hash 需要计算的哈希值,ahash/phash/dhash"""
    print_function_info()
    try:
        image_pil = Image.open(imagefile)
        resize_image_pil = image_pil.resize((8, 8))
        if mode_hash == 'ahash':
            calc_hash = imagehash.average_hash(resize_image_pil)  # 均值哈希
        elif mode_hash == 'phash':
            calc_hash = imagehash.phash(resize_image_pil)  # 感知哈希
        elif mode_hash == 'dhash':
            calc_hash = imagehash.dhash(resize_image_pil)  # 差异哈希
        else:
            calc_hash = None
    except:
        calc_hash = None

    # 转为01二进制字符串，方便存储和读取
    if calc_hash:
        calc_hash_str = hash_numpy2str(calc_hash)
    else:
        calc_hash_str = None

    return calc_hash_str


def compare_image(image_data_dict, mode_ahash=True, mode_phash=True, mode_dhash=True, mode_ssim=True):
    """传入数据字典，对比其中的图片，查找相似项"""
    print_function_info()
    similar_group_list = []  # 相似组列表 [(源文件1,源文件2), (...)...]
    all_image_list = list(image_data_dict.keys())
    key_index = -1  # 当前检查的图的索引
    # 提取检查的图
    for key in all_image_list[:-1]:  # 最后一张图不需要进行检查
        key_index += 1
        key_origin = image_data_dict[key]['origin_path']
        key_ahash = image_data_dict[key]['ahash']
        key_phash = image_data_dict[key]['phash']
        key_dhash = image_data_dict[key]['dhash']
        # 提取对比的图
        for compare in all_image_list[key_index + 1:]:  # 其后的所有图都需要对比
            compare_origin = image_data_dict[compare]['origin_path']
            compare_ahash = image_data_dict[compare]['ahash']
            compare_phash = image_data_dict[compare]['phash']
            compare_dhash = image_data_dict[compare]['dhash']
            # 检查是否为同一源文件，是则跳过
            if key_origin == compare_origin:
                continue
            # 对比hash值
            similar_check = []  # 存放bool值，通过内部True的个数来判断是否为相似
            if mode_ahash:
                if key_ahash is not None and compare_ahash is not None:
                    dist_ahash = calc_two_hash_str_hamming_distance(key_ahash, compare_ahash)
                    if dist_ahash <= mode_ahash:
                        similar_check.append(True)
                    else:
                        similar_check.append(False)
                else:
                    similar_check.append(False)

            if mode_phash:
                if key_phash is not None and compare_phash is not None:
                    dist_phash = calc_two_hash_str_hamming_distance(key_phash, compare_phash)
                    if dist_phash <= mode_phash:
                        similar_check.append(True)
                    else:
                        similar_check.append(False)
                else:
                    similar_check.append(False)

            if mode_dhash:
                if key_dhash is not None and compare_dhash is not None:
                    dist_dhash = calc_two_hash_str_hamming_distance(key_dhash, compare_dhash)
                    if dist_dhash <= mode_dhash:
                        similar_check.append(True)
                    else:
                        similar_check.append(False)
                else:
                    similar_check.append(False)
            # 判断hash对比结果
            if True in similar_check:
                if mode_ssim:
                    ssim = calc_images_ssim(key_origin, compare_origin)
                    if ssim >= mode_ssim:
                        similar_group_list.append((key_origin, compare_origin))
                else:
                    similar_group_list.append((key_origin, compare_origin))
    # 处理相似组列表（去重、合并交集项）
    final_similar_group_list = merge_intersecting_tuples(similar_group_list)

    return final_similar_group_list


def save_similar_result(data_list, date_dict):
    """保存相似项结果到xlsx"""
    print_function_info()
    if not data_list:
        return
    if not os.path.exists(history_back_dir):
        os.mkdir(history_back_dir)
    current_time = time.strftime('%Y-%m-%d %H_%M_%S', time.localtime())
    xlsx_name = f'{history_back_dir}/查重结果 {current_time}.xlsx'
    # 新建空xlsx
    wb = Workbook()
    wb.save(xlsx_name)
    # 读取xlsx
    wb = load_workbook(xlsx_name)
    ws = wb.active
    # 写入数据
    index_number = 0
    index_row = 1
    ws['A1'].value = '相似组'
    ws['B1'].value = '文件路径'
    ws['C1'].value = '文件类型'
    ws['D1'].value = '文件大小/MB'
    ws['E1'].value = '内部图片数量'
    for data_set in data_list:
        index_number += 1
        ws[f'A{index_row + 1}'].value = f'相似组 {index_number}'
        for i in data_set:
            ws[f'B{index_row + 1}'].value = i
            ws[f'B{index_row + 1}'].hyperlink = i
            ws[f'B{index_row + 1}'].font = styles.Font(underline="single", color="0000FF")
            ws[f'C{index_row + 1}'].value = date_dict[i]['filetype']
            ws[f'D{index_row + 1}'].value = date_dict[i]['filesize'] / 1024 / 1024
            ws[f'E{index_row + 1}'].value = date_dict[i]['image_number']
            index_row += 1
    # 保存
    wb.save(filename=xlsx_name)


def hash_numpy2str(hash_numpy):
    """将哈希值的numpy数组(imagehash.hash)转换为二进制字符串"""
    if not hash_numpy:
        return None
    if type(hash_numpy) is imagehash.ImageHash:
        hash_numpy = hash_numpy.hash
    hash_str = ''
    for row in hash_numpy:
        for col in row:
            if col:
                hash_str += '1'
            else:
                hash_str += '0'

    return hash_str


def calc_two_hash_str_similar(hash_str1, hash_str2):
    """计算两个二进制字符串哈希值的相似度"""
    hash_int1 = int(hash_str1, 2)
    hash_int2 = int(hash_str2, 2)
    # 使用异或操作计算差异位数
    diff_bits = bin(hash_int1 ^ hash_int2).count('1')
    # 计算相似性
    similarity = 1 - diff_bits / len(hash_str1)

    return similarity


def calc_two_hash_str_hamming_distance(hash_str1, hash_str2):
    """计算两个二进制字符串哈希值的汉明距离"""
    hamming_distance = sum(ch1 != ch2 for ch1, ch2 in zip(hash_str1, hash_str2))

    return hamming_distance


def check_hash_cache():
    """检查哈希值缓存文件"""
    # 检查数据库是否存在
    conn = sqlite3.connect(cache_filename)
    cursor = conn.cursor()
    cursor.execute(
        'CREATE TABLE IF NOT EXISTS Table_Hash (path TEXT Primary KEY, filesize INTEGER, ahash TEXT, phash TEXT, dhash TEXT)')
    # 读取数据库记录
    cursor.execute("SELECT * FROM Table_Hash")
    data = cursor.fetchall()
    columns_name = [i[0] for i in cursor.description]
    data_origin = [dict(zip(columns_name, row)) for row in data]
    # 提取需要的数据，并进行检查
    file_cache_data = {}  # {源文件路径:{'filesize':int, 'ahash':'str', ...}...}
    for group_dict in data_origin:
        path = group_dict['path']
        filesize = group_dict['filesize']
        # 检查路径是否存在
        if os.path.exists(path):
            # 检查路径对应的文件大小是否正确
            if os.path.isdir(path):
                real_filesize = get_folder_size(path)
            else:
                real_filesize = os.path.getsize(path)
            if real_filesize == int(filesize):
                file_cache_data[path] = {}
                for key, value in group_dict.items():
                    if key != 'path':
                        if value == 'None':  # 从数据库取得的None值为字符串类型，而非python的None类型
                            value = None
                        file_cache_data[path][key] = value

    return file_cache_data


def update_hash_cache(new_file_cache_data):
    """更新哈希值缓存文件
    传参：特定格式的字典，{源文件路径:{'filesize':int, 'ahash':'str', ...}...}"""
    conn = sqlite3.connect(cache_filename)
    cursor = conn.cursor()
    # 获取路径列中的数据
    cursor.execute("SELECT path FROM Table_Hash")
    data = cursor.fetchall()
    exist_data_path = []
    for group_tuple in data:
        exist_data_path += group_tuple
    # 更新或插入新数据
    for path_key in new_file_cache_data:
        if path_key in exist_data_path:  # 若存在则更新
            for option, value in new_file_cache_data[path_key].items():
                cursor.execute(f'UPDATE Table_Hash SET {option} = "{str(value)}" WHERE path = "{path_key}"')
        else:  # 不存在则新增
            cursor.execute(f'INSERT INTO Table_Hash (path) VALUES ("{path_key}")')
            for option, value in new_file_cache_data[path_key].items():
                cursor.execute(f'UPDATE Table_Hash SET {option} = "{str(value)}" WHERE path = "{path_key}"')
    # 更新并关闭数据库
    cursor.close()
    conn.commit()
    conn.close()
